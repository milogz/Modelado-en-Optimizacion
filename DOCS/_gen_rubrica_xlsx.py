# -*- coding: utf-8 -*-
"""Genera el archivo XLSX editable de la rúbrica del curso IIND2501."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Estilos ──────────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
mod_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
mod_font = Font(name="Calibri", bold=True, size=11, color="1F3864")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
level_fills = {
    1: PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # rojo claro
    2: PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid"),  # naranja claro
    3: PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # verde claro
    4: PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),  # azul claro
}

# ── Datos ────────────────────────────────────────────────────────────────
objetivos_intermedios = [
    # (id, descripción, OG, lecciones, actividades,
    #  criterio_1_insuficiente, criterio_2_en_desarrollo, criterio_3_competente, criterio_4_sobresaliente)

    # --- Módulo 1 ---
    ("OI-1.1",
     "Reconocer los elementos canónicos de un problema de optimización (variables de decisión, F.O., restricciones) a partir de un enunciado en lenguaje natural.",
     "OG-1, OG-2", "MO_3-0, MO_3-3", "E1-P1, E1-P2",
     "No identifica variables de decisión, F.O. o restricciones, o los confunde gravemente.",
     "Identifica variables y F.O. pero omite restricciones relevantes o las define de forma ambigua.",
     "Identifica correctamente todos los elementos (conjuntos, parámetros, variables, restricciones, F.O.) con notación clara.",
     "Además justifica por qué eligió esas variables, discute supuestos implícitos y propone extensiones."),

    ("OI-1.2",
     "Representar gráficamente un LP de 2 variables: región factible, gradiente, isoclinas, punto óptimo; usar la representación para razonar sobre cambios en parámetros.",
     "OG-2, OG-4", "MO_3-3, Anexo 1", "E1-P1 a–d",
     "No genera la gráfica o ésta no refleja las restricciones del problema.",
     "Grafica la mayoría de las restricciones pero no identifica correctamente la región factible, el gradiente o el punto óptimo.",
     "Grafica todas las restricciones, sombrea la región factible, traza el gradiente y las isoclinas, e identifica el punto óptimo.",
     "Además usa la gráfica para razonar sobre cambios paramétricos (escenarios) y explica restricciones activas/inactivas."),

    ("OI-1.3",
     "Formular modelos LP continuos con notación indexada (conjuntos, parámetros, variables, restricciones, F.O.) para problemas de transporte/distribución.",
     "OG-2", "MO_3-3, MO_3-4", "E1-P2",
     "Formulación ausente o con errores estructurales (dimensiones incompatibles, restricciones faltantes).",
     "Estructura correcta pero errores en índices, dominios o signos que harían el modelo incorrecto.",
     "Formulación general correcta con conjuntos, parámetros y variables bien definidos; restricciones y F.O. consistentes.",
     "Formulación impecable y generalizable; discute alternativas de modelado y justifica la elección de índices."),

    ("OI-1.4",
     "Realizar análisis de sensibilidad: precio sombra, rango de variación de un parámetro (RHS) y graficar la F.O. paramétrica.",
     "OG-3, OG-4", "MO_3-5", "E1-P2 f–g",
     "No reporta precios sombra ni rangos, o los reporta incorrectamente.",
     "Reporta precios sombra pero no los rangos, o los interpreta de forma genérica sin anclaje al problema.",
     "Reporta precios sombra y rangos correctamente; grafica la F.O. paramétrica y explica las pendientes.",
     "Interpreta precios sombra en contexto del problema, explica discontinuidades de la gráfica y justifica la selección del recurso."),

    ("OI-1.5",
     "Formular modelos MILP con variables binarias y condiciones lógicas (if-then, at-least-k, exclusive-or, condicionales compuestas).",
     "OG-2", "MO_3-7", "E1-P3, E1-P4",
     "No introduce variables binarias o las condiciones lógicas están mal traducidas a restricciones lineales.",
     "Modela algunas condiciones lógicas correctamente pero omite o traduce incorrectamente al menos una.",
     "Todas las condiciones lógicas correctamente traducidas a restricciones lineales con las M adecuadas.",
     "Compara diferentes formalizaciones, tightens las M, o discute la fortaleza de la formulación."),

    ("OI-1.6",
     "Implementar y resolver un LP/MILP en Solver de Excel y en Python-PuLP, documentando el código y verificando coherencia.",
     "OG-3", "MO_3-3, MO_3-7", "E1-P2 c+d, E1-P3 b, E1-P4 c",
     "El código no ejecuta o produce resultados incorrectos; no implementa en ambas plataformas cuando se solicita.",
     "Código funcional en una plataforma pero con errores en la otra, o sin documentación.",
     "Ambas implementaciones funcionales, documentadas y con resultados coherentes entre sí.",
     "Código limpio, modular, bien documentado; verifica automáticamente coherencia entre Solver y PuLP."),

    ("OI-1.7",
     "Presentar resultados de un modelo de optimización de manera organizada (tablas, gráficas, diagramas) y redactar conclusiones que reflejen comprensión.",
     "OG-4", "Transversal", "E1: todos los literales 'c'/'e'",
     "Resultados crudos sin organización, o ausencia de conclusiones.",
     "Presenta resultados en tablas/gráficas pero conclusiones superficiales o no conectan con el problema.",
     "Resultados bien organizados con tablas, gráficas pertinentes y conclusiones que reflejan comprensión.",
     "Comunicación de nivel profesional: visualizaciones limpias, narración de hallazgos, recomendaciones y limitaciones."),

    # --- Módulo 2 ---
    ("OI-2.1",
     "Formular un problema de optimización no lineal y distinguirlo de un LP; reconocer cuándo los supuestos de linealidad no se cumplen.",
     "OG-1", "MO_3-1, MO_3-6", "E2-P1 a",
     "No formula el problema o lo trata erróneamente como lineal.",
     "Formula correctamente la F.O. o las restricciones pero no ambas; no menciona la no-linealidad.",
     "Formulación completa con F.O. no lineal y restricciones de dominio; señala qué supuestos de LP no se cumplen.",
     "Analiza convexidad/concavidad, discute óptimos locales/globales y conecta con la estrategia de solución."),

    ("OI-2.2",
     "Diseñar e implementar un algoritmo de búsqueda local y analizar su convergencia y dependencia del punto inicial.",
     "OG-3", "MO_3-1", "E2-P1 b–e",
     "El algoritmo no converge o no implementa correctamente el mecanismo de vecindad/factibilidad.",
     "Implementa el radar y la búsqueda pero no analiza dependencia del punto inicial o reporta resultados incompletos.",
     "Código funcional para las 4 funciones; tabla completa para 3 puntos iniciales; conclusión sobre óptimos locales vs. globales.",
     "Propone mejoras (multi-start, λ adaptativo), o genera visualización del espacio de búsqueda."),

    ("OI-2.3",
     "Convertir un LP entre formatos (canónico → estándar → matricial) y resolver paso a paso con el Método Simplex.",
     "OG-2, OG-3", "MO_3-4", "E2-P2 a–e",
     "Errores graves en las conversiones de formato o iteraciones del Simplex (pivot incorrecto, no termina).",
     "Conversiones correctas pero iteraciones con errores aritméticos o de selección de variable entrante/saliente.",
     "Conversiones y todas las iteraciones correctas, con chequeo de optimalidad en cada paso.",
     "Interpreta geométricamente cada iteración, o discute degeneracía/ciclaje/regla de selección alternativa."),

    ("OI-2.4",
     "Derivar el problema dual de un LP, resolverlo computacionalmente e interpretar variables duales (precios sombra) y holgura complementaria.",
     "OG-2, OG-4", "MO_3-5", "E2-P2 f–i",
     "No formula el dual o lo formula incorrectamente.",
     "Dual correcto y resuelto, pero interpretación genérica ('precio sombra' sin contexto).",
     "Dual correcto, tabla primal-dual completa, holgura complementaria verificada, interpretación anclada al problema.",
     "Discute implicaciones gerenciales de los precios sombra, propone qué recurso expandir y conecta con análisis de sensibilidad."),

    ("OI-2.5",
     "Formular un TSP como MILP y comparar la solución exacta con heurísticas constructivas usando métricas de gap, distancia y tiempo.",
     "OG-1, OG-3, OG-4", "MO_3-7, MO_3-1", "E2-P3",
     "No implementa las heurísticas o no formula el MILP del TSP.",
     "Implementa ambas aproximaciones pero la comparación es superficial o carece de métricas cuantitativas.",
     "MILP y heurísticas correctos; tabla comparativa con distancia, gap% y tiempo; discusión de trade-off.",
     "Visualiza rutas, discute complejidad computacional, o explora variantes (2-opt, diferentes criterios de inserción)."),

    ("OI-2.6",
     "Diferenciar problemas de decisión (analítica prescriptiva) de problemas de estimación y reconocer el rol de la optimización en ambos.",
     "OG-1", "MO_3-2, MO_3-0", "Magistral semanas 10–12",
     "Confunde los dos paradigmas o no logra articular la diferencia.",
     "Distingue vagamente ('uno es para predecir, otro para decidir') sin ejemplos ni relación con la optimización.",
     "Diferencia clara con ejemplos del curso; identifica el rol de la optimización en ambos (F.O. ajuste vs. F.O. decisión).",
     "Plantea escenarios híbridos, conecta con regresión/gradient descent y el hilo narrativo del curso."),

    # --- Módulo 3 ---
    ("OI-3.1",
     "Definir los elementos de un modelo de Programación Dinámica (estados, acciones, transición, costo, Bellman) a partir de un enunciado.",
     "OG-1, OG-2", "MO_3-8", "E2-P4 d",
     "No define estados, acciones o ecuación de Bellman, o los confunde con los de un LP.",
     "Define parcialmente los elementos pero la función de transición o la ecuación de Bellman tiene errores.",
     "Todos los elementos correctamente definidos y consistentes con el enunciado.",
     "Generaliza la formulación, discute estado terminal vs. horizonte infinito, o conecta con inducción del valor."),

    ("OI-3.2",
     "Implementar y evaluar políticas de decisión definidas por reglas, registrando trayectorias de estados y costo acumulado.",
     "OG-3", "MO_3-8", "E2-P4 e–f",
     "Las funciones de política no compilan o producen trayectorias infactibles.",
     "Políticas funcionales pero sin registrar correctamente la trayectoria completa o sin calcular el costo acumulado.",
     "Ambas políticas implementadas correctamente, trayectorias y costos totales reportados, comparación con el óptimo.",
     "Visualiza estados a lo largo del horizonte, compara con métricas adicionales, o propone políticas alternativas."),

    ("OI-3.3",
     "Implementar Policy Evaluation para calcular el valor de cada estado bajo una política dada y contrastar con el costo forward.",
     "OG-3", "MO_3-8", "E2-P4 g",
     "No implementa Policy Evaluation o la implementación es incorrecta.",
     "Implementación funcional pero no contrasta los valores con los obtenidos en la simulación forward.",
     "Policy Evaluation correcta, contraste explícito con valores forward, conclusión sobre consistencia.",
     "Discute convergencia, interpreta los valores en contexto del problema, o conecta con ecuaciones de valor en RL."),

    ("OI-3.4",
     "Aplicar Backwards Induction (ecuación de Bellman) para encontrar la política óptima y compararla con la solución MIP.",
     "OG-3, OG-4", "MO_3-8", "E2-P4 b vs. h",
     "No implementa backwards induction o la política óptima resultante es incorrecta.",
     "Backwards induction funcional pero no compara con la solución MIP, o no verifica equivalencia.",
     "Política óptima vía Bellman correcta, coincide con MIP; discusión de por qué ambos dan el mismo resultado.",
     "Discute ventajas/desventajas de cada paradigma (escalabilidad, curva de la dimensionalidad), o extiende a incertidumbre."),

    ("OI-3.5",
     "Contrastar paradigmas de solución (LP/MILP, heurísticas, DP) y argumentar cuándo cada enfoque es apropiado.",
     "OG-1, OG-4", "MO_3-0, MO_3-8", "E2-P3, E2-P4",
     "No logra articular diferencias entre LP/MILP, heurísticas y DP.",
     "Lista los métodos pero no fundamenta cuándo usar cada uno basándose en características del problema.",
     "Argumenta criterios de selección (linealidad, secuencialidad, tamaño, garantía de optimalidad) con ejemplos del curso.",
     "Elabora un marco de decisión ('árbol de selección'), conecta con la práctica profesional y menciona extensiones."),
]

modulos = [
    ("Módulo 1 — Fundamentos de Optimización y LP (Semanas 1–8 · Entrenamiento 1)", 0, 7),
    ("Módulo 2 — Más allá de lo Lineal: Búsqueda, Simplex, Dualidad (Semanas 9–12 · Entrenamiento 2)", 7, 13),
    ("Módulo 3 — Decisiones Secuenciales: DP y Políticas (Semanas 13–16 · Entrenamiento 2)", 13, 18),
]

# ── Hoja 1: Rúbrica completa ────────────────────────────────────────────
ws = wb.active
ws.title = "Rúbrica"

# Anchos de columna
col_widths = [10, 45, 12, 18, 20, 40, 40, 40, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Encabezados
headers = ["ID", "Objetivo Intermedio", "OG", "Lección(es)", "Actividad(es)",
           "1 – Insuficiente", "2 – En desarrollo", "3 – Competente", "4 – Sobresaliente"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

ws.freeze_panes = "A2"
row = 2

for mod_name, start, end in modulos:
    # Fila de módulo
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=mod_name)
    cell.font = mod_font
    cell.fill = mod_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = thin_border
    ws.row_dimensions[row].height = 28
    row += 1

    for oi in objetivos_intermedios[start:end]:
        oi_id, desc, og, lecs, acts, c1, c2, c3, c4 = oi
        vals = [oi_id, desc, og, lecs, acts, c1, c2, c3, c4]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = wrap
            cell.border = thin_border
            # Color de fondo para columnas de nivel
            if c == 6:
                cell.fill = level_fills[1]
            elif c == 7:
                cell.fill = level_fills[2]
            elif c == 8:
                cell.fill = level_fills[3]
            elif c == 9:
                cell.fill = level_fills[4]
        ws.row_dimensions[row].height = 72
        row += 1

# ── Hoja 2: Objetivos generales ─────────────────────────────────────────
ws2 = wb.create_sheet("Objetivos Generales")
ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 90

og_headers = ["ID", "Objetivo General (Syllabus)"]
for c, h in enumerate(og_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

og_data = [
    ("OG-1", "Identificar situaciones susceptibles de ser mejoradas mediante técnicas de optimización; específicamente, problemas de decisión y de estimación."),
    ("OG-2", "Formular problemas de optimización, representando situaciones reales de decisión a través de modelos matemáticos de programación lineal."),
    ("OG-3", "Implementar y resolver modelos de optimización mediante herramientas existentes (PuLP, Solver) así como mediante desarrollos propios."),
    ("OG-4", "Analizar, interpretar y comunicar apropiadamente los resultados de un modelo de optimización."),
]
for r, (oid, desc) in enumerate(og_data, 2):
    ws2.cell(row=r, column=1, value=oid).border = thin_border
    c = ws2.cell(row=r, column=2, value=desc)
    c.alignment = wrap
    c.border = thin_border

# ── Hoja 3: Matriz de cobertura OG × OI ─────────────────────────────────
ws3 = wb.create_sheet("Cobertura OG-OI")
ws3.column_dimensions["A"].width = 10
for i in range(2, 6):
    ws3.column_dimensions[get_column_letter(i)].width = 16

cov_headers = ["OI", "OG-1 Identificar", "OG-2 Formular", "OG-3 Implementar", "OG-4 Comunicar"]
for c, h in enumerate(cov_headers, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap
    cell.border = thin_border

coverage_map = {
    "OI-1.1": [True, True, False, False],
    "OI-1.2": [False, True, False, True],
    "OI-1.3": [False, True, False, False],
    "OI-1.4": [False, False, True, True],
    "OI-1.5": [False, True, False, False],
    "OI-1.6": [False, False, True, False],
    "OI-1.7": [False, False, False, True],
    "OI-2.1": [True, False, False, False],
    "OI-2.2": [False, False, True, False],
    "OI-2.3": [False, True, True, False],
    "OI-2.4": [False, True, False, True],
    "OI-2.5": [True, False, True, True],
    "OI-2.6": [True, False, False, False],
    "OI-3.1": [True, True, False, False],
    "OI-3.2": [False, False, True, False],
    "OI-3.3": [False, False, True, False],
    "OI-3.4": [False, False, True, True],
    "OI-3.5": [True, False, False, True],
}

check_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for r, oi in enumerate(objetivos_intermedios, 2):
    oi_id = oi[0]
    ws3.cell(row=r, column=1, value=oi_id).border = thin_border
    for c, covered in enumerate(coverage_map[oi_id], 2):
        cell = ws3.cell(row=r, column=c, value="✓" if covered else "")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        if covered:
            cell.fill = check_fill

# ── Guardar ──────────────────────────────────────────────────────────────
out = r"DOCS\IIND2501_Rubrica_Objetivos.xlsx"
wb.save(out)
print(f"✅ Archivo generado: {out}")
